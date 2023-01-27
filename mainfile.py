import nltk
from os import listdir
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords

stopword=set()
for entry in listdir('StopWords/'):
    stopword.update(set(words.strip() for words in open('StopWords/'+entry)))

def result_to_excel(res, loc):
    outb = load_workbook('Output.xlsx')
    out = outb.active
    for i in range(len(res)):
        out[chr(67+i)+str(loc)] = res[i]
        outb.save('Output.xlsx')

def word_count(word):
    stopword.update(stopwords.words('english'))
    count=0
    for aword in word:
        if aword in stopword or (not aword and aword[0] in {'?', '!', ',', '.' }): continue
        count+=1
    return count

def syllable_count_per_word(word):
    compxword=0
    syllablecount=0
    for aword in word:
        count=0
        for char in aword:
            if char in 'AEIOUaeiou':
                count+=1
        if len(aword)>1 and aword[-2]=='e':
            if aword[-1]=='s' or aword[-1]=='d':
                count-=1
        if count>2: compxword+=1
        syllablecount+=count
    return [compxword, syllablecount/len(word)]

def personal_pronouns(word):
    count, charcount = 0, 0
    for aword,pos in nltk.pos_tag(word):
        charcount+=len(aword)
        if aword in {'I','we','my','ours','us','US'}:
            continue
        if pos=='PRP':
            count+=1
    return [count, charcount/len(word)]

def analysis_of_Readability(complex_word, word, sentence):
    avg_sent_len = len(word)/len(sentence)
    per_cmpx_word = complex_word/len(word)
    return [avg_sent_len, per_cmpx_word, 0.4 * (avg_sent_len + per_cmpx_word)]

def sentimental_Analysis(word):
    posFile = set(words.strip() for words in open('MasterDictionary/positive-words.txt'))
    negFile = set(words.strip() for words in open('MasterDictionary/negative-words.txt'))
    pos, neg, totstopinword = 0, 0, 0
    
    for aword in word:
        if aword in stopword: 
            totstopinword+=1
            continue
        if aword in posFile: pos+=1
        if aword in negFile: neg-=1
    neg=neg*(-1)
    polarity_score = (pos-neg)/((pos+neg)+0.000001)
    subjectivity_score = (pos+neg)/((len(word)-totstopinword)+0.000001)
    return [pos, neg, polarity_score, subjectivity_score]

def main(data):
    ans, word, sentence = [], [], []

    for i in data:
        l=sent_tokenize(i)
        sentence.extend(l)
        word.extend(word_tokenize(i))
    
    ans.extend(sentimental_Analysis(word))
    val = syllable_count_per_word(word)
    ans.extend(analysis_of_Readability(val[0], word, sentence))
    ans.extend([len(word)/len(sentence), val[0], word_count(word), val[1]])
    ans.extend(personal_pronouns(word))
    return ans

if __name__ == "__main__":
    ip = load_workbook('input.xlsx').active
    for i in range(2, len(ip['B'])):
        url = ip['B'+str(i)].value
        # url = 'https://insights.blackcoffer.com/will-ai-replace-us-or-work-with-us/'
        r = requests.get(url ,headers={"User-Agent": "XY"})
        data = []
        soup = BeautifulSoup(r.content, 'html.parser')
        for j in soup.find_all({'title', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p'}):
            data.append(j.get_text())
        result_to_excel(main(data), i)
        